"""
Generate Survey123 XLSForm workbooks from spec definitions.

Usage:
    python -m field_survey.build_xlsforms [--output-dir field_survey/xlsform]
    [--choices-dir field_survey/choices]

Requires: openpyxl
"""
import csv
import os

import openpyxl


# ── Helpers ─────────────────────────────────────────────────────────────

def _likert(name, label, low="1 - Strongly disagree", high="5 - Strongly agree"):
    return {
        "type": "select_one likert5",
        "name": name,
        "label": label,
        "hint": f"{low} ... {high}",
        "required": "yes",
    }


def _select(name, label, list_name, **kwargs):
    return {"type": f"select_one {list_name}", "name": name, "label": label, "required": "yes", **kwargs}


def _multi(name, label, list_name, **kwargs):
    return {"type": f"select_multiple {list_name}", "name": name, "label": label, "required": "yes", **kwargs}


def _text(name, label, **kwargs):
    return {"type": "text", "name": name, "label": label, **kwargs}


def _group(name, label):
    return {"type": "begin_group", "name": name, "label": label}


def _end():
    return {"type": "end_group"}


def _note(label):
    return {"type": "note", "name": f"note_{abs(hash(label)) % 10000}", "label": label}


# ── Inline choice lists ────────────────────────────────────────────────

INLINE_CHOICES = {
    "yes_no": [("yes", "Yes"), ("no", "No")],
    "yes_no_dk": [("yes", "Yes"), ("no", "No"), ("dont_know", "Don't know")],
    "yes_no_pnts": [("yes", "Yes"), ("no", "No"), ("pnts", "Prefer not to say")],
    "consent_gate": [("yes", "Yes, I consent"), ("no", "No, I do not consent")],
    "likert5": [("1", "1"), ("2", "2"), ("3", "3"), ("4", "4"), ("5", "5")],
    "gender": [("man", "Man"), ("woman", "Woman"), ("nonbinary", "Non-binary"), ("pnts", "Prefer not to say")],
    "age_range": [("18_24", "18-24"), ("25_34", "25-34"), ("35_44", "35-44"), ("45_54", "45-54"), ("55_64", "55-64"), ("65_plus", "65+")],
    "age_range_student": [("18_21", "18-21"), ("22_25", "22-25"), ("26_30", "26-30"), ("30_plus", "30+")],
    "household_size": [("1", "1"), ("2_3", "2-3"), ("4_5", "4-5"), ("6_plus", "6+")],
    "years_neighbourhood": [("lt1", "Less than 1 year"), ("1_3", "1-3 years"), ("3_5", "3-5 years"), ("5_10", "5-10 years"), ("10_20", "10-20 years"), ("20_plus", "20+ years")],
    "tenure": [("rent", "Rent"), ("own", "Own"), ("social", "Social housing"), ("other", "Other")],
    "unit_type": [("house", "House"), ("apartment", "Apartment"), ("rooming", "Rooming house"), ("above_shop", "Above-shop unit"), ("other", "Other")],
    "rent_range_res": [("lt800", "Less than $800"), ("800_1200", "$800-$1,200"), ("1200_1600", "$1,200-$1,600"), ("1600_2000", "$1,600-$2,000"), ("2000_plus", "$2,000+"), ("na", "N/A")],
    "rent_increase_pct": [("lt2", "Less than 2%"), ("2_5", "2-5%"), ("5_10", "5-10%"), ("gt10", "More than 10%"), ("dk", "Don't know")],
    "changes_3yr": [("major", "Major changes"), ("some", "Some changes"), ("none", "No change"), ("new", "New to area")],
    "sentiment": [("positive", "Positive"), ("negative", "Negative"), ("mixed", "Mixed"), ("neutral", "Neutral")],
    "dev_impact": [("positive", "Positive"), ("negative", "Negative"), ("mixed", "Mixed"), ("no_opinion", "No opinion")],
    "concern": [("affordability", "Affordability"), ("safety", "Safety"), ("noise", "Noise"), ("cleanliness", "Cleanliness"), ("traffic", "Traffic"), ("development", "Development"), ("loss_character", "Loss of character"), ("other", "Other")],
    "asset": [("community", "Community"), ("diversity", "Diversity"), ("walkability", "Walkability"), ("food", "Food"), ("markets", "Markets"), ("culture", "Culture"), ("greenspace", "Greenspace"), ("affordability", "Affordability"), ("other", "Other")],
    "business_type": [("restaurant", "Restaurant"), ("cafe", "Cafe"), ("retail", "Retail"), ("market_vendor", "Market vendor"), ("service", "Service"), ("bar", "Bar"), ("grocery", "Grocery"), ("other", "Other")],
    "years_operating": [("lt1", "Less than 1 year"), ("1_3", "1-3 years"), ("3_5", "3-5 years"), ("5_10", "5-10 years"), ("10_20", "10-20 years"), ("20_plus", "20+ years")],
    "ownership": [("owner", "Owner-operated"), ("franchise", "Franchise"), ("family", "Family business"), ("partnership", "Partnership"), ("other", "Other")],
    "employees": [("1_2", "1-2"), ("3_5", "3-5"), ("6_10", "6-10"), ("11_20", "11-20"), ("20_plus", "20+")],
    "lease_type": [("month", "Month-to-month"), ("short", "Short-term (< 3 years)"), ("long", "Long-term (3+ years)"), ("own", "Own the building"), ("pnts", "Prefer not to say")],
    "rent_range_biz": [("lt2k", "Less than $2,000"), ("2_4k", "$2,000-$4,000"), ("4_6k", "$4,000-$6,000"), ("6_10k", "$6,000-$10,000"), ("10k_plus", "$10,000+"), ("pnts", "Prefer not to say")],
    "rent_change": [("inc_lot", "Increased a lot"), ("inc_some", "Increased somewhat"), ("stable", "Stable"), ("decreased", "Decreased"), ("new_lease", "New lease"), ("pnts", "Prefer not to say")],
    "revenue_trend": [("growing", "Growing"), ("stable", "Stable"), ("declining", "Declining"), ("pnts", "Prefer not to say")],
    "customer_change": [("tourists", "More tourists"), ("students", "More students"), ("locals", "More locals"), ("no_change", "No change"), ("mixed", "Mixed")],
    "traffic_trend": [("increasing", "Increasing"), ("stable", "Stable"), ("decreasing", "Decreasing"), ("seasonal", "Seasonal")],
    "competition": [("more", "More competition"), ("less", "Less competition"), ("same", "Same"), ("different", "Different type of competition")],
    "closure_count": [("1_2", "1-2"), ("3_5", "3-5"), ("5_plus", "5+")],
    "threat": [("rent", "Rising rent"), ("clientele", "Changing clientele"), ("competition", "Competition"), ("development", "Development"), ("parking", "Parking"), ("crime", "Crime"), ("regulation", "Regulation"), ("none", "None"), ("other", "Other")],
    "opportunity": [("tourism", "Tourism"), ("new_residents", "New residents"), ("events", "Events"), ("online", "Online presence"), ("community", "Community support"), ("other", "Other")],
    "patio": [("yes", "Yes"), ("no", "No"), ("denied", "Applied but denied"), ("na", "Not applicable")],
    "heritage_dk": [("yes", "Yes"), ("no", "No"), ("dk", "Don't know")],
    "bia": [("yes", "Yes"), ("no", "No"), ("dk", "Don't know")],
    "community_inv": [("bia", "BIA"), ("market_events", "Market events"), ("neighbourhood_assoc", "Neighbourhood association"), ("none", "None"), ("other", "Other")],
    "plans_3yr": [("stay", "Stay"), ("expand", "Expand"), ("downsize", "Downsize"), ("relocate", "Relocate"), ("close", "Close"), ("uncertain", "Uncertain")],
    "connection": [("live", "Live here"), ("work", "Work here"), ("visiting", "Visiting"), ("shopping", "Shopping"), ("passing", "Passing through"), ("student", "Student nearby")],
    "visit_freq": [("daily", "Daily"), ("few_week", "Few times a week"), ("weekly", "Weekly"), ("monthly", "Monthly"), ("first", "First time"), ("rarely", "Rarely")],
    "visit_reason": [("food", "Food"), ("shopping", "Shopping"), ("restaurant_bar", "Restaurant/Bar"), ("work", "Work"), ("live", "Live here"), ("exploring", "Exploring"), ("meeting", "Meeting someone"), ("other", "Other")],
    "transport": [("walk", "Walk"), ("bike", "Bike"), ("ttc", "TTC"), ("car", "Car"), ("rideshare", "Rideshare")],
    "transport_student": [("walk", "Walk"), ("bike", "Bike"), ("ttc", "TTC"), ("car", "Car"), ("other", "Other")],
    "time_spent": [("lt30", "Less than 30 min"), ("30_60", "30 min - 1 hour"), ("1_2hr", "1-2 hours"), ("2_4hr", "2-4 hours"), ("4_plus", "4+ hours")],
    "money_spent": [("0", "$0"), ("lt20", "Less than $20"), ("20_50", "$20-$50"), ("50_100", "$50-$100"), ("100_plus", "$100+")],
    "changes_yn": [("yes", "Yes"), ("no", "No"), ("dk", "Don't visit enough to know")],
    "recommend": [("yes", "Yes"), ("no", "No"), ("maybe", "Maybe")],
    "improve_intercept": [("cleanliness", "Cleanliness"), ("safety", "Safety"), ("seating", "More seating"), ("washrooms", "Public washrooms"), ("less_traffic", "Less traffic"), ("greenery", "More greenery"), ("nothing", "Nothing"), ("other", "Other")],
    "student_status": [("undergrad", "Undergraduate"), ("graduate", "Graduate"), ("postdoc", "Post-doc"), ("staff", "Staff")],
    "faculty": [("arts_sci", "Arts & Science"), ("engineering", "Engineering"), ("architecture", "Architecture"), ("planning", "Planning"), ("social_work", "Social Work"), ("other", "Other")],
    "lives_kenso": [("yes", "Yes"), ("no", "No"), ("used_to", "Used to")],
    "housing_student": [("on_campus", "On-campus"), ("rent_nearby", "Rent nearby"), ("rent_elsewhere", "Rent elsewhere"), ("family", "With family"), ("other", "Other")],
    "visit_freq_student": [("daily", "Daily"), ("few_week", "Few times a week"), ("weekly", "Weekly"), ("monthly", "Monthly"), ("rarely", "Rarely"), ("never", "Never")],
    "visit_reason_student": [("food", "Food"), ("bars", "Bars"), ("shopping", "Shopping"), ("vintage", "Vintage stores"), ("exploring", "Exploring"), ("friends", "Friends live there"), ("live", "Live there"), ("never", "Never visit")],
    "would_live": [("yes", "Yes"), ("already", "Already do"), ("no", "No"), ("used_to", "Used to")],
    "barrier": [("rent", "Rent too high"), ("safety", "Safety"), ("far", "Too far from campus"), ("noise", "Noise"), ("quality", "Housing quality"), ("none", "No barrier"), ("other", "Other")],
    "gentrif_aware": [("yes", "Yes"), ("somewhat", "Somewhat"), ("no", "No")],
    "gentrif_opinion": [("positive", "Positive"), ("negative", "Negative"), ("mixed", "Mixed"), ("no_opinion", "No opinion")],
    "student_impact": [("drive_up", "Students drive up rents"), ("priced_out", "Students are also priced out"), ("no_impact", "No impact"), ("dk", "Don't know")],
    "improve_student": [("affordability", "Affordability"), ("safety", "Safety"), ("cleanliness", "Cleanliness"), ("student_spaces", "More student spaces"), ("transit", "Better transit"), ("nothing", "Nothing"), ("other", "Other")],
    "campus_loc": [("robarts", "Robarts Library"), ("sidney_smith", "Sidney Smith"), ("hart_house", "Hart House"), ("spadina_college", "Spadina & College"), ("spadina_dundas", "Spadina & Dundas"), ("other", "Other")],
    "contact_method": [("email", "Email"), ("phone", "Phone"), ("either", "Either")],
    "language_pref": [("english", "English"), ("french", "French"), ("mandarin", "Mandarin"), ("cantonese", "Cantonese"), ("portuguese", "Portuguese"), ("spanish", "Spanish"), ("other", "Other")],
    "surveyor": [(f"surveyor_{i}", f"Surveyor {i}") for i in range(1, 7)],
}

# ── Form 1: Resident Survey ────────────────────────────────────────────

RESIDENT_SURVEY = [
    _note("**Kensington Market Resident Survey**\n\nThis survey is anonymous. Your responses will be aggregated at the block level."),
    _group("metadata", "Survey Information"),
    _select("surveyor_id", "Surveyor", "surveyor"),
    {"type": "dateTime", "name": "survey_timestamp", "label": "Date/Time", "default": "now()"},
    _select("block_location", "Block Location", "block_location", appearance="search"),
    _select("informed_consent", "Do you consent to participate in this anonymous survey?", "consent_gate"),
    {"type": "hidden", "name": "form_version", "label": "Form Version", "default": "1.0"},
    _end(),
    _group("demographics", "About You"),
    _select("age_range", "Age range", "age_range"),
    _select("gender", "Gender", "gender"),
    _select("household_size_range", "Household size", "household_size"),
    _select("years_in_neighbourhood", "How long have you lived in this neighbourhood?", "years_neighbourhood"),
    _text("primary_language", "What is your primary language at home?", required="no"),
    _end(),
    _group("housing", "Housing"),
    _select("tenure_type", "Do you rent or own?", "tenure"),
    _select("unit_type", "What type of unit do you live in?", "unit_type"),
    _select("monthly_rent_range", "Monthly rent range", "rent_range_res", relevant="${tenure_type} != 'own'"),
    _select("rent_increase_last_year", "Has your rent increased in the last year?", "yes_no_dk", relevant="${tenure_type} != 'own'"),
    _select("rent_increase_pct", "By approximately how much?", "rent_increase_pct", relevant="${rent_increase_last_year} = 'yes'"),
    _likert("fear_of_displacement", "How worried are you about being displaced from your home?", "1 - Not at all worried", "5 - Extremely worried"),
    _select("received_eviction_notice", "Have you received an eviction notice in the past 2 years?", "yes_no_pnts"),
    _end(),
    _group("neighbourhood", "Your Neighbourhood"),
    _likert("neighbourhood_satisfaction", "Overall satisfaction with the neighbourhood", "1 - Very dissatisfied", "5 - Very satisfied"),
    _multi("biggest_concern", "What are your biggest concerns? (select all that apply)", "concern"),
    _multi("biggest_asset", "What do you value most about the neighbourhood? (select all)", "asset"),
    _likert("perceived_safety_day", "How safe do you feel during the day?", "1 - Very unsafe", "5 - Very safe"),
    _likert("perceived_safety_night", "How safe do you feel at night?", "1 - Very unsafe", "5 - Very safe"),
    _select("noticed_changes_3yr", "Have you noticed changes in the last 3 years?", "changes_3yr"),
    _text("change_description", "Describe the changes you've noticed", relevant="${noticed_changes_3yr} != 'none' and ${noticed_changes_3yr} != 'new'", required="no"),
    _select("change_sentiment", "Overall, are these changes...", "sentiment", relevant="${noticed_changes_3yr} != 'none' and ${noticed_changes_3yr} != 'new'"),
    _end(),
    _group("services", "Access to Services"),
    _likert("access_grocery", "Ease of access to grocery stores", "1 - Very difficult", "5 - Very easy"),
    _likert("access_healthcare", "Ease of access to healthcare", "1 - Very difficult", "5 - Very easy"),
    _likert("access_transit", "Ease of access to public transit", "1 - Very difficult", "5 - Very easy"),
    _likert("access_greenspace", "Ease of access to parks/green space", "1 - Very difficult", "5 - Very easy"),
    _text("missing_services", "What service or amenity is missing from this neighbourhood?", required="no"),
    _end(),
    _group("gentrification", "Neighbourhood Change"),
    _select("aware_of_development", "Are you aware of new development in the area?", "yes_no"),
    _select("development_impact", "How do you feel about the development?", "dev_impact", relevant="${aware_of_development} = 'yes'"),
    _select("business_closures_noticed", "Have you noticed business closures recently?", "yes_no"),
    _text("closure_names", "Which businesses closed?", relevant="${business_closures_noticed} = 'yes'", required="no"),
    _likert("community_belonging", "How strongly do you feel you belong to this community?", "1 - Not at all", "5 - Very strongly"),
    _end(),
    _group("close", "Thank You"),
    _select("consent_followup", "Would you be willing to be contacted for a follow-up?", "yes_no"),
    _text("additional_comments", "Any other comments?", required="no"),
    _end(),
]

# ── Form 2: Business Owner Survey ──────────────────────────────────────

BUSINESS_SURVEY = [
    _note("**Kensington Market Business Survey**\n\nYour business may be identified in published results. No personal names will be recorded."),
    _group("metadata", "Survey Information"),
    _select("surveyor_id", "Surveyor", "surveyor"),
    {"type": "dateTime", "name": "survey_timestamp", "label": "Date/Time", "default": "now()"},
    _select("business_name", "Business Name", "business_name", appearance="search"),
    _select("street_address", "Street Address", "street_address", appearance="search"),
    {"type": "geopoint", "name": "gps_location", "label": "GPS Location"},
    _select("informed_consent", "I understand my business may be identified. No personal names will be published.", "consent_gate"),
    {"type": "hidden", "name": "form_version", "label": "Form Version", "default": "1.0"},
    _end(),
    _group("profile", "Business Profile"),
    _select("business_type", "Type of business", "business_type"),
    _select("years_operating", "Years operating at this location", "years_operating"),
    _select("ownership_type", "Ownership type", "ownership"),
    _select("num_employees_range", "Number of employees", "employees"),
    _select("is_original_business", "Was this business founded at this location?", "yes_no"),
    _end(),
    _group("economics", "Rent & Economics"),
    _select("lease_type", "Lease type", "lease_type"),
    _select("monthly_rent_range", "Monthly rent", "rent_range_biz"),
    _select("rent_change_3yr", "How has rent changed in 3 years?", "rent_change"),
    _select("revenue_trend_3yr", "Revenue trend over 3 years", "revenue_trend"),
    _likert("financial_viability", "Financial viability of business", "1 - At risk of closing", "5 - Very secure"),
    _end(),
    _group("change", "Neighbourhood Change"),
    _select("customer_base_change", "How has your customer base changed?", "customer_change"),
    _select("foot_traffic_trend", "Foot traffic trend", "traffic_trend"),
    _select("competition_change", "How has competition changed?", "competition"),
    _select("nearby_closures_noticed", "Noticed nearby business closures?", "yes_no"),
    _select("closure_count_estimate", "How many closures?", "closure_count", relevant="${nearby_closures_noticed} = 'yes'"),
    _likert("gentrification_impact", "Impact of gentrification on your business", "1 - Very negative", "5 - Very positive"),
    _multi("biggest_threat", "Biggest threats to your business (select all)", "threat"),
    _multi("biggest_opportunity", "Biggest opportunities (select all)", "opportunity"),
    _end(),
    _group("operations", "Operations"),
    _select("patio_program", "Participate in CafeTO patio program?", "patio"),
    _likert("accessibility_rating", "How accessible is your business?", "1 - Not accessible", "5 - Fully accessible"),
    _select("delivery_apps", "Do you use delivery apps?", "yes_no"),
    _select("heritage_building", "Is this a heritage building?", "heritage_dk"),
    _end(),
    _group("community", "Community"),
    _select("belongs_to_bia", "Member of a BIA?", "bia"),
    _multi("community_involvement", "Community involvement (select all)", "community_inv"),
    _likert("neighbourhood_satisfaction", "Neighbourhood satisfaction", "1 - Very dissatisfied", "5 - Very satisfied"),
    _select("plans_next_3yr", "Plans for next 3 years", "plans_3yr"),
    _end(),
    _group("close", "Thank You"),
    _text("additional_comments", "Any other comments?", required="no"),
    _select("consent_followup", "Willing to be contacted for follow-up?", "yes_no"),
    _end(),
]

# ── Form 3: Street Intercept Survey ────────────────────────────────────

INTERCEPT_SURVEY = [
    _note("**Kensington Market Visitor Survey** (3-4 minutes)\n\nThis survey is anonymous."),
    _group("metadata", "Survey Information"),
    _select("surveyor_id", "Surveyor", "surveyor"),
    {"type": "dateTime", "name": "survey_timestamp", "label": "Date/Time", "default": "now()"},
    _select("block_location", "Block Location", "block_location", appearance="search"),
    _select("informed_consent", "Do you consent to this anonymous survey?", "consent_gate"),
    {"type": "hidden", "name": "form_version", "label": "Form Version", "default": "1.0"},
    _end(),
    _group("who", "About You"),
    _select("age_range", "Age range", "age_range"),
    _select("connection_to_area", "Connection to Kensington", "connection"),
    _select("visit_frequency", "How often do you visit?", "visit_freq"),
    _end(),
    _group("experience", "Your Visit Today"),
    _multi("reason_for_visit", "Why are you here today? (select all)", "visit_reason"),
    _select("how_arrived", "How did you get here?", "transport"),
    _select("time_spent_today", "Time spent here today", "time_spent"),
    _select("money_spent_today", "Money spent today", "money_spent"),
    _end(),
    _group("perception", "Your Impression"),
    _likert("overall_impression", "Overall impression of Kensington", "1 - Very negative", "5 - Very positive"),
    _likert("perceived_safety", "How safe do you feel?", "1 - Very unsafe", "5 - Very safe"),
    _likert("cleanliness", "Cleanliness", "1 - Very dirty", "5 - Very clean"),
    _likert("accessibility", "Accessibility", "1 - Not accessible", "5 - Very accessible"),
    _likert("vibrancy", "How lively does it feel?", "1 - Dead", "5 - Very lively"),
    _end(),
    _group("change_section", "Change"),
    _select("noticed_changes", "Have you noticed changes?", "changes_yn"),
    _select("change_sentiment", "Are the changes...", "sentiment", relevant="${noticed_changes} = 'yes'"),
    _text("one_word_kensington", "One word to describe Kensington?"),
    _end(),
    _group("close", "Thank You"),
    _select("would_recommend", "Would you recommend visiting Kensington?", "recommend"),
    _multi("what_would_improve", "What would improve Kensington? (select all)", "improve_intercept"),
    _end(),
]

# ── Form 4: Student Survey ─────────────────────────────────────────────

STUDENT_SURVEY = [
    _note("**UofT Student Survey — Kensington Market** (4-5 minutes)\n\nThis survey is anonymous."),
    _group("metadata", "Survey Information"),
    _select("surveyor_id", "Surveyor", "surveyor"),
    {"type": "dateTime", "name": "survey_timestamp", "label": "Date/Time", "default": "now()"},
    _select("campus_location", "Where on campus are you now?", "campus_loc"),
    _select("informed_consent", "Do you consent to this anonymous survey?", "consent_gate"),
    {"type": "hidden", "name": "form_version", "label": "Form Version", "default": "1.0"},
    _end(),
    _group("profile", "About You"),
    _select("age_range", "Age range", "age_range_student"),
    _select("student_status", "Student status", "student_status"),
    _select("faculty", "Faculty", "faculty"),
    _select("lives_in_kensington_area", "Do you live in/near Kensington?", "lives_kenso"),
    _select("housing_type", "Housing type", "housing_student"),
    _select("monthly_rent_range", "Monthly rent", "rent_range_res"),
    _end(),
    _group("kensington", "Kensington Market"),
    _select("visit_frequency", "How often do you visit Kensington?", "visit_freq_student"),
    _multi("reason_for_visit", "Why do you visit? (select all)", "visit_reason_student"),
    _select("how_arrived", "How do you usually get there?", "transport_student", relevant="${visit_frequency} != 'never'"),
    _select("money_spent_typical", "Typical spend per visit", "money_spent", relevant="${visit_frequency} != 'never'"),
    _end(),
    _group("perception", "Your Perception"),
    _likert("overall_impression", "Overall impression of Kensington", "1 - Very negative", "5 - Very positive"),
    _likert("perceived_safety", "How safe does Kensington feel?", "1 - Very unsafe", "5 - Very safe"),
    _likert("sense_of_community", "Sense of community", "1 - No community feel", "5 - Strong community"),
    _likert("authenticity", "How authentic does it feel?", "1 - Commercialized", "5 - Authentic"),
    _likert("affordability_perception", "How affordable is Kensington?", "1 - Too expensive", "5 - Very affordable"),
    _end(),
    _group("housing_gentrif", "Housing & Gentrification"),
    _select("would_live_in_kensington", "Would you live in Kensington?", "would_live"),
    _multi("barrier_to_living_there", "What prevents you? (select all)", "barrier", relevant="${would_live_in_kensington} = 'no'"),
    _select("aware_of_gentrification", "Aware of gentrification in Kensington?", "gentrif_aware"),
    _select("gentrification_opinion", "Your opinion on gentrification there", "gentrif_opinion"),
    _select("student_housing_impact", "Impact of students on Kensington housing", "student_impact"),
    _end(),
    _group("close", "Thank You"),
    _text("one_word_kensington", "One word to describe Kensington?"),
    _multi("what_would_improve", "What would improve Kensington? (select all)", "improve_student"),
    _end(),
]

# ── Form 5: Follow-up Contact ──────────────────────────────────────────

CONTACT_FORM = [
    _note("**Follow-up Contact Form**\n\nThis form is NOT linked to any survey response."),
    _text("random_code", "Random code from card"),
    _text("name", "Name (optional)", required="no"),
    _select("contact_method", "Preferred contact method", "contact_method"),
    _text("contact_info", "Email or phone number"),
    _select("preferred_language", "Preferred language", "language_pref"),
]


# ── XLSForm writer ─────────────────────────────────────────────────────

def _write_xlsform(form_fields, form_title, outfile, choices_dir=None):
    """Write an XLSForm .xlsx workbook."""
    wb = openpyxl.Workbook()

    # survey sheet
    ws = wb.active
    ws.title = "survey"
    headers = ["type", "name", "label", "hint", "required", "relevant", "appearance", "default"]
    ws.append(headers)

    for field in form_fields:
        row = [
            field.get("type", ""),
            field.get("name", ""),
            field.get("label", ""),
            field.get("hint", ""),
            field.get("required", ""),
            field.get("relevant", ""),
            field.get("appearance", ""),
            field.get("default", ""),
        ]
        ws.append(row)

    # choices sheet
    ws_choices = wb.create_sheet("choices")
    ws_choices.append(["list_name", "name", "label"])

    # Collect used choice lists
    used_lists = set()
    for field in form_fields:
        ftype = field.get("type", "")
        for prefix in ("select_one ", "select_multiple "):
            if ftype.startswith(prefix):
                used_lists.add(ftype[len(prefix):])

    # Write inline choices
    for list_name in sorted(used_lists):
        if list_name in INLINE_CHOICES:
            for name, label in INLINE_CHOICES[list_name]:
                ws_choices.append([list_name, name, label])

    # Append external choice lists from CSV
    if choices_dir:
        for csv_file in ["business_names.csv", "addresses.csv", "block_locations.csv"]:
            csv_path = os.path.join(choices_dir, csv_file)
            if os.path.exists(csv_path):
                with open(csv_path, "r", encoding="utf-8") as f:
                    reader = csv.DictReader(f)
                    for row in reader:
                        if row.get("list_name") in used_lists:
                            ws_choices.append([row["list_name"], row["name"], row["label"]])

    # settings sheet
    ws_settings = wb.create_sheet("settings")
    ws_settings.append(["form_title", "form_id", "version", "style"])
    form_id = form_title.lower().replace(" ", "_").replace("-", "_")
    ws_settings.append([form_title, form_id, "1.0", "theme-grid"])

    wb.save(outfile)


def main():
    import argparse

    parser = argparse.ArgumentParser(description="Generate Survey123 XLSForm workbooks")
    parser.add_argument("--output-dir", default="field_survey/xlsform")
    parser.add_argument("--choices-dir", default="field_survey/choices")
    args = parser.parse_args()

    os.makedirs(args.output_dir, exist_ok=True)

    forms = [
        (RESIDENT_SURVEY, "Kensington Resident Survey", "resident_survey.xlsx"),
        (BUSINESS_SURVEY, "Kensington Business Survey", "business_survey.xlsx"),
        (INTERCEPT_SURVEY, "Kensington Intercept Survey", "intercept_survey.xlsx"),
        (STUDENT_SURVEY, "UofT Student Survey", "student_survey.xlsx"),
        (CONTACT_FORM, "Follow-up Contact", "contact_followup.xlsx"),
    ]

    for fields, title, filename in forms:
        outpath = os.path.join(args.output_dir, filename)
        _write_xlsform(fields, title, outpath, args.choices_dir)
        print(f"Created {outpath} ({len(fields)} fields)")


if __name__ == "__main__":
    main()

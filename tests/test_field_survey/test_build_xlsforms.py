import os
import pytest
import openpyxl

from field_survey.build_xlsforms import (
    RESIDENT_SURVEY,
    BUSINESS_SURVEY,
    INTERCEPT_SURVEY,
    STUDENT_SURVEY,
    CONTACT_FORM,
    INLINE_CHOICES,
    _write_xlsform,
)


ALL_FORMS = [
    ("resident", RESIDENT_SURVEY),
    ("business", BUSINESS_SURVEY),
    ("intercept", INTERCEPT_SURVEY),
    ("student", STUDENT_SURVEY),
    ("contact", CONTACT_FORM),
]


def test_all_forms_have_informed_consent():
    """Every survey form (except contact) must have an informed_consent field."""
    for name, fields in ALL_FORMS:
        if name == "contact":
            continue
        consent_fields = [f for f in fields if f.get("name") == "informed_consent"]
        assert len(consent_fields) == 1, f"{name} form missing informed_consent"
        assert "consent_gate" in consent_fields[0]["type"], f"{name} consent must use consent_gate list"


def test_all_forms_have_form_version():
    """Every survey form (except contact) must have a hidden form_version field."""
    for name, fields in ALL_FORMS:
        if name == "contact":
            continue
        version_fields = [f for f in fields if f.get("name") == "form_version"]
        assert len(version_fields) == 1, f"{name} form missing form_version"
        assert version_fields[0]["type"] == "hidden", f"{name} form_version must be hidden"
        assert version_fields[0]["default"] == "1.0"


def test_resident_survey_has_skip_logic():
    """Resident survey should have relevant (skip logic) on rent questions."""
    rent_fields = [f for f in RESIDENT_SURVEY if f.get("name") == "monthly_rent_range"]
    assert len(rent_fields) == 1
    assert "relevant" in rent_fields[0]
    assert "tenure_type" in rent_fields[0]["relevant"]


def test_inline_choices_cover_all_used_lists():
    """Every choice list referenced by forms must exist in INLINE_CHOICES or be an external CSV list."""
    external_lists = {"business_name", "street_address", "block_location"}
    missing = []
    for name, fields in ALL_FORMS:
        for field in fields:
            ftype = field.get("type", "")
            for prefix in ("select_one ", "select_multiple "):
                if ftype.startswith(prefix):
                    list_name = ftype[len(prefix):]
                    if list_name not in INLINE_CHOICES and list_name not in external_lists:
                        missing.append(f"{name}:{field.get('name')}→{list_name}")
    assert missing == [], f"Missing choice lists: {missing}"


def test_write_xlsform_creates_valid_workbook(tmp_path):
    """Generated workbook must have survey, choices, and settings sheets."""
    outfile = str(tmp_path / "test_form.xlsx")
    _write_xlsform(INTERCEPT_SURVEY, "Test Form", outfile)

    assert os.path.exists(outfile)
    wb = openpyxl.load_workbook(outfile)
    assert "survey" in wb.sheetnames
    assert "choices" in wb.sheetnames
    assert "settings" in wb.sheetnames

    # survey sheet has headers + data rows
    ws = wb["survey"]
    headers = [cell.value for cell in ws[1]]
    assert "type" in headers
    assert "name" in headers
    assert "label" in headers
    assert ws.max_row > 1  # has data beyond headers

    # choices sheet has entries
    ws_choices = wb["choices"]
    assert ws_choices.max_row > 1

    # settings sheet has form_title
    ws_settings = wb["settings"]
    assert ws_settings.cell(2, 1).value == "Test Form"

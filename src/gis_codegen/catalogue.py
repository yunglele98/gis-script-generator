"""
gis_codegen.catalogue

Reads the Kensington Market map catalogue Excel file and generates one
PyQGIS or ArcPy script per map entry (filtered to status=have|partial,
Vector layers).

CLI:
    gis-catalogue --input catalogue.xlsx --output-dir ./maps/
    gis-catalogue --input catalogue.xlsx --platform arcpy --output-dir ./maps_arcpy/
    gis-catalogue --input catalogue.xlsx --schema schema.json --op buffer
"""

import json
import os
import sys
import argparse
from pathlib import Path
from datetime import datetime

try:
    import openpyxl
except ImportError:
    print("[ERROR] openpyxl is required: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

from gis_codegen.generator import (
    safe_var,
    VALID_OPERATIONS,
    _pyqgis_op_blocks,
    _arcpy_op_blocks,
    generate_qgs,
    generate_pyt,
)

# ---------------------------------------------------------------------------
# Catalogue loader
# ---------------------------------------------------------------------------

VALID_STATUSES    = {"have", "partial"}
RASTER_ONLY_TYPES = {"Raster", "Table"}   # skip if *only* these (no Vector component)

_NUMERIC_TYPES = {
    "integer", "bigint", "smallint", "double precision",
    "numeric", "real", "float4", "float8",
}
_TEXT_TYPES = {"text", "character varying", "character", "varchar"}


def _best_field(layer_info: dict | None, numeric: bool = True) -> str:
    """
    Return the most suitable column name from a schema layer dict.

    Skips the geometry column and primary keys. Falls back to 'value'
    when layer_info is absent or no suitable column is found.
    """
    if not layer_info:
        return "value"
    geom_col = layer_info.get("geometry", {}).get("column", "geom")
    pks      = set(layer_info.get("primary_keys", []))
    want     = _NUMERIC_TYPES if numeric else _TEXT_TYPES
    for col in layer_info.get("columns", []):
        name  = col["name"]
        dtype = col.get("data_type", "")
        if name == geom_col or name in pks:
            continue
        if dtype in want:
            return name
    return "value"


def load_schema(path: str) -> dict:
    """
    Load a schema JSON produced by  gis-codegen --save-schema.
    Returns a dict keyed by table name for fast lookup.
    """
    data   = json.loads(Path(path).read_text(encoding="utf-8"))
    lookup = {}
    for layer in data.get("layers", []):
        lookup[layer["table"]] = layer
    return lookup


def load_catalogue(path: str) -> list[dict]:
    """
    Read Catalogue sheet, return map dicts for Vector layers with
    status=have or status=partial.
    """
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        print(f"[ERROR] Cannot open catalogue file '{path}': {exc}", file=sys.stderr)
        sys.exit(1)

    if "Catalogue" not in wb.sheetnames:
        print(
            f"[ERROR] Worksheet 'Catalogue' not found in '{path}'.\n"
            f"        Available sheets: {wb.sheetnames}",
            file=sys.stderr,
        )
        sys.exit(1)

    ws   = wb["Catalogue"]
    rows = list(ws.iter_rows(values_only=True))
    headers = rows[0]

    maps = []
    for row in rows[1:]:
        d = dict(zip(headers, row))
        status     = (d.get("status") or "").strip().lower()
        layer_type = (d.get("spatial_layer_type") or "").strip()
        if status in VALID_STATUSES and "Vector" in layer_type:
            maps.append(d)
    return maps


# ---------------------------------------------------------------------------
# Symbology block builders
# ---------------------------------------------------------------------------

def _graduated_block(var: str, field_hint: str, ramp: str = "YlOrRd",
                     n_classes: int = 5) -> list[str]:
    return [
        f'    from qgis.core import (',
        f'        QgsGraduatedSymbolRenderer, QgsClassificationQuantile,',
        f'        QgsColorBrewerColorRamp,',
        f'    )',
        f'    GRAD_FIELD_{var} = "{field_hint}"  # TODO: verify field name',
        f'    _rend_{var} = QgsGraduatedSymbolRenderer(GRAD_FIELD_{var})',
        f'    _rend_{var}.setClassificationMethod(QgsClassificationQuantile())',
        f'    _rend_{var}.updateClasses(lyr_{var}, {n_classes})',
        f'    _rend_{var}.updateColorRamp(QgsColorBrewerColorRamp("{ramp}", {n_classes}))',
        f'    lyr_{var}.setRenderer(_rend_{var})',
        f'    lyr_{var}.triggerRepaint()',
        f'    print(f"  Renderer: graduated on \'{{GRAD_FIELD_{var}}}\', {n_classes} classes")',
    ]


def _categorized_block(var: str, field_hint: str, classification: str) -> list[str]:
    return [
        f'    from qgis.core import (',
        f'        QgsCategorizedSymbolRenderer, QgsRendererCategory,',
        f'        QgsSymbol, QgsStyle,',
        f'    )',
        f'    CAT_FIELD_{var} = "{field_hint}"  # TODO: verify field name',
        f'    # Classification scheme: {classification or "see catalogue"}',
        f'    # Build categories automatically from unique values:',
        f'    _cats_{var} = []',
        f'    for _val in lyr_{var}.uniqueValues(',
        f'        lyr_{var}.fields().indexFromName(CAT_FIELD_{var})',
        f'    ):',
        f'        _sym = QgsSymbol.defaultSymbol(lyr_{var}.geometryType())',
        f'        _cats_{var}.append(QgsRendererCategory(_val, _sym, str(_val)))',
        f'    _rend_{var} = QgsCategorizedSymbolRenderer(CAT_FIELD_{var}, _cats_{var})',
        f'    _rend_{var}.updateColorRamp(',
        f'        QgsStyle.defaultStyle().colorRamp("Paired"),',
        f'    )',
        f'    lyr_{var}.setRenderer(_rend_{var})',
        f'    lyr_{var}.triggerRepaint()',
        f'    print(f"  Renderer: categorized on \'{{CAT_FIELD_{var}}}\'")',
    ]


def _network_line_block(var: str) -> list[str]:
    return [
        f'    from qgis.core import (',
        f'        QgsCategorizedSymbolRenderer, QgsRendererCategory,',
        f'        QgsLineSymbol, QgsSymbol,',
        f'    )',
        f'    NET_FIELD_{var} = "route_type"  # TODO: verify field name (e.g. route_type, highway)',
        f'    _cats_{var} = []',
        f'    for _val in lyr_{var}.uniqueValues(',
        f'        lyr_{var}.fields().indexFromName(NET_FIELD_{var})',
        f'    ):',
        f'        _sym = QgsLineSymbol.createSimple({{"width": "0.5"}})',
        f'        _cats_{var}.append(QgsRendererCategory(_val, _sym, str(_val)))',
        f'    _rend_{var} = QgsCategorizedSymbolRenderer(NET_FIELD_{var}, _cats_{var})',
        f'    lyr_{var}.setRenderer(_rend_{var})',
        f'    lyr_{var}.triggerRepaint()',
        f'    print(f"  Renderer: network categorized on \'{{NET_FIELD_{var}}}\'")',
    ]


def _heatmap_block(var: str) -> list[str]:
    return [
        f'    from qgis.core import QgsHeatmapRenderer, QgsStyle',
        f'    _heat_{var} = QgsHeatmapRenderer()',
        f'    _heat_{var}.setRadius(15)',
        f'    _heat_{var}.setMaximumValue(0)  # 0 = auto',
        f'    _heat_{var}.setColorRamp(',
        f'        QgsStyle.defaultStyle().colorRamp("Reds"),',
        f'    )',
        f'    lyr_{var}.setRenderer(_heat_{var})',
        f'    lyr_{var}.triggerRepaint()',
        f'    print("  Renderer: heatmap density")',
    ]


def _points_polygons_block(var: str) -> list[str]:
    return [
        f'    from qgis.core import QgsSingleSymbolRenderer, QgsSymbol',
        f'    _sym_{var} = QgsSymbol.defaultSymbol(lyr_{var}.geometryType())',
        f'    lyr_{var}.setRenderer(QgsSingleSymbolRenderer(_sym_{var}))',
        f'    lyr_{var}.triggerRepaint()',
        f'    print("  Renderer: single symbol (points/polygons) — customise in Layer Properties")',
    ]


def _symbology_block(var: str, m: dict) -> list[str]:
    """Dispatch to the appropriate renderer block based on symbology_type."""
    stype   = (m.get("symbology_type") or "").lower()
    classif = (m.get("classification") or "")
    lines   = [
        f'    # --- Symbology: {m.get("symbology_type", "(unknown)")} ---',
    ]

    if "heatmap" in stype or "densité" in stype:
        lines += _heatmap_block(var)

    elif "réseau" in stype or "network" in stype:
        lines += _network_line_block(var)

    elif ("catégoriel" in stype or "catégorie" in stype) and "choroplèthe" not in stype:
        lines += _categorized_block(var, "type", classif)

    elif "choroplèthe" in stype or "dégradé" in stype or "gradué" in stype:
        if "catégoriel" in stype:
            lines += _categorized_block(var, "type", classif)
        else:
            lines += _graduated_block(var, classif or "value")

    elif "points" in stype or "polygones" in stype:
        lines += _points_polygons_block(var)

    else:
        lines += [
            f'    # TODO: configure renderer',
            f'    # Symbology type: {m.get("symbology_type")}',
            f'    # Use Layer Properties → Symbology in QGIS to configure interactively.',
        ]

    return lines


# ---------------------------------------------------------------------------
# ArcPy renderer block builders
# ---------------------------------------------------------------------------

def _arcpy_graduated_block(var: str, field_hint: str, n_classes: int = 5) -> list[str]:
    return [
        f'    GRAD_FIELD_{var} = "{field_hint}"  # TODO: verify field name',
        f'    sym_{var} = lyr_{var}.symbology',
        f'    sym_{var}.updateRenderer("GraduatedColorsRenderer")',
        f'    sym_{var}.renderer.classificationField = GRAD_FIELD_{var}',
        f'    sym_{var}.renderer.breakCount = {n_classes}',
        f'    # TODO: sym_{var}.renderer.colorRamp = aprx.listColorRamps("Oranges ({n_classes} Classes)")[0]',
        f'    lyr_{var}.symbology = sym_{var}',
        f'    print(f"  Renderer: graduated on \'{{GRAD_FIELD_{var}}}\', {n_classes} classes")',
    ]


def _arcpy_categorized_block(var: str, field_hint: str, classification: str) -> list[str]:
    return [
        f'    CAT_FIELD_{var} = "{field_hint}"  # TODO: verify field name',
        f'    # Classification scheme: {classification or "see catalogue"}',
        f'    sym_{var} = lyr_{var}.symbology',
        f'    sym_{var}.updateRenderer("UniqueValueRenderer")',
        f'    sym_{var}.renderer.fields = [CAT_FIELD_{var}]',
        f'    lyr_{var}.symbology = sym_{var}',
        f'    print(f"  Renderer: unique values on \'{{CAT_FIELD_{var}}}\'")',
    ]


def _arcpy_heatmap_block(var: str) -> list[str]:
    return [
        f'    # Requires ArcGIS Pro 3.x',
        f'    sym_{var} = lyr_{var}.symbology',
        f'    sym_{var}.updateRenderer("HeatMapRenderer")',
        f'    # TODO: configure radius and color scheme in Layer Properties',
        f'    lyr_{var}.symbology = sym_{var}',
        f'    print("  Renderer: heat map density")',
    ]


def _arcpy_network_line_block(var: str) -> list[str]:
    return [
        f'    NET_FIELD_{var} = "route_type"  # TODO: verify field name (e.g. route_type, highway)',
        f'    sym_{var} = lyr_{var}.symbology',
        f'    sym_{var}.updateRenderer("UniqueValueRenderer")',
        f'    sym_{var}.renderer.fields = [NET_FIELD_{var}]',
        f'    lyr_{var}.symbology = sym_{var}',
        f'    print(f"  Renderer: network unique values on \'{{NET_FIELD_{var}}}\'")',
    ]


def _arcpy_points_polygons_block(var: str) -> list[str]:
    return [
        f'    sym_{var} = lyr_{var}.symbology',
        f'    sym_{var}.updateRenderer("SimpleRenderer")',
        f'    # TODO: customise symbol colour/size in Layer Properties',
        f'    lyr_{var}.symbology = sym_{var}',
        f'    print("  Renderer: simple symbol (points/polygons)")',
    ]


def _arcpy_symbology_block(var: str, m: dict) -> list[str]:
    """Dispatch to the appropriate ArcPy renderer block based on symbology_type."""
    stype   = (m.get("symbology_type") or "").lower()
    classif = (m.get("classification") or "")
    lines   = [
        f'    # --- Symbology: {m.get("symbology_type", "(unknown)")} ---',
    ]

    if "heatmap" in stype or "densité" in stype:
        lines += _arcpy_heatmap_block(var)

    elif "réseau" in stype or "network" in stype:
        lines += _arcpy_network_line_block(var)

    elif ("catégoriel" in stype or "catégorie" in stype) and "choroplèthe" not in stype:
        lines += _arcpy_categorized_block(var, "type", classif)

    elif "choroplèthe" in stype or "dégradé" in stype or "gradué" in stype:
        if "catégoriel" in stype:
            lines += _arcpy_categorized_block(var, "type", classif)
        else:
            lines += _arcpy_graduated_block(var, classif or "value")

    elif "points" in stype or "polygones" in stype:
        lines += _arcpy_points_polygons_block(var)

    else:
        lines += [
            f'    # TODO: configure renderer',
            f'    # Symbology type: {m.get("symbology_type")}',
            f'    # Use Layer Properties → Symbology in ArcGIS Pro.',
        ]

    return lines


# ---------------------------------------------------------------------------
# Per-map PyQGIS script generator
# ---------------------------------------------------------------------------

def generate_map_pyqgis(m: dict, db_config: dict,
                        ops: list | None = None,
                        layer_info: dict | None = None) -> str:
    map_id     = m.get("map_id", "M??")
    title      = m.get("title") or m.get("short_name", "")
    short_name = m.get("short_name", "layer")
    theme      = m.get("theme", "")
    subtheme   = m.get("subtheme", "")
    objective  = m.get("objective", "")
    questions  = m.get("key_questions", "")
    indicators = m.get("key_indicators", "")
    scale      = m.get("study_scale", "")
    unit       = m.get("unit_of_analysis", "")
    classif    = m.get("classification", "")
    sources    = m.get("data_sources", "")
    vintage    = m.get("data_vintage", "")
    processing = m.get("processing_steps", "")
    symbology  = m.get("symbology_type", "")
    deliverable= m.get("deliverable_format", "")
    validation = m.get("validation_checks", "")
    risks      = m.get("risks_limitations", "")
    status     = m.get("status", "")
    owner      = m.get("owner", "")
    layer_type = m.get("spatial_layer_type", "Vector")

    var  = safe_var(short_name)
    host = db_config["host"]
    port = db_config["port"]
    db   = db_config["dbname"]
    user = db_config["user"]

    has_raster = "Raster" in layer_type

    lines = [
        f'"""',
        f'Map ID    : {map_id}',
        f'Title     : {title}',
        f'Theme     : {theme} > {subtheme}',
        f'Objective : {objective}',
        f'Questions : {questions}',
        f'Indicators: {indicators}',
        f'Scale     : {scale}  |  Unit: {unit}',
        f'Symbology : {symbology}',
        f'Sources   : {sources}  [{vintage}]',
        f'Processing: {processing}',
        f'Deliverable: {deliverable}',
        f'Validation: {validation}',
        f'Risks     : {risks}',
        f'Status    : {status}  |  Owner: {owner}',
        f'Generated : {datetime.now().strftime("%Y-%m-%d %H:%M")}',
        f'"""',
        f'',
        f'import os',
        f'',
        f'from qgis.core import (',
        f'    QgsApplication, QgsDataSourceUri, QgsVectorLayer, QgsProject,',
        f')',
        f'',
        f'qgs = QgsApplication([], False)',
        f'qgs.initQgis()',
        f'',
        f'DB_HOST     = "{host}"',
        f'DB_PORT     = "{port}"',
        f'DB_NAME     = "{db}"',
        f'DB_USER     = "{user}"',
        f'DB_PASSWORD = os.environ["PGPASSWORD"]',
        f'',
        f'# {"=" * 66}',
        f'# {map_id} — {short_name}',
        f'# {title}',
        f'# {"=" * 66}',
        f'',
        f'uri_{var} = QgsDataSourceUri()',
        f'uri_{var}.setConnection(DB_HOST, DB_PORT, DB_NAME, DB_USER, DB_PASSWORD)',
        f'uri_{var}.setDataSource("public", "{short_name}", "geom", "", "")',
        f'',
        f'lyr_{var} = QgsVectorLayer(uri_{var}.uri(False), "{short_name}", "postgres")',
        f'',
        f'if not lyr_{var}.isValid():',
        f'    print(f"[ERROR] \'{short_name}\' failed to load.")',
        f'else:',
        f'    QgsProject.instance().addMapLayer(lyr_{var})',
        f'    print(f"[OK] {short_name}: {{lyr_{var}.featureCount()}} features")',
        f'    print(f"  CRS: {{lyr_{var}.crs().authid()}}")',
        f'    print(f"  Fields: {{[f.name() for f in lyr_{var}.fields()]}}")',
        f'',
    ]

    # Symbology — enrich field hint from schema if catalogue classification is blank
    m_sym = m
    if layer_info and not m.get("classification"):
        stype = (m.get("symbology_type") or "").lower()
        want_numeric = any(k in stype for k in ("choroplèthe", "dégradé", "gradué", "heatmap", "densité"))
        best = _best_field(layer_info, numeric=want_numeric)
        if best != "value":
            m_sym = {**m, "classification": best}
    lines.extend(_symbology_block(var, m_sym))
    lines.append(f'')

    # Operation blocks (--op)
    if ops:
        columns = layer_info.get("columns", []) if layer_info else []
        lines.extend(_pyqgis_op_blocks(var, short_name, columns, set(ops)))

    # Raster note
    if has_raster:
        lines += [
            f'    # --- Raster component ---',
            f'    # This layer also has a raster component ({layer_type}).',
            f'    # TODO: load raster layer via QgsRasterLayer and add to project.',
            f'    # raster_path = r"TODO: path to raster file"',
            f'    # lyr_{var}_raster = QgsRasterLayer(raster_path, "{short_name}_raster")',
            f'    # QgsProject.instance().addMapLayer(lyr_{var}_raster)',
            f'',
        ]

    # Validation checks
    if validation:
        checks = [c.strip() for c in str(validation).split(",")]
        lines += [
            f'    # --- Validation checks ---',
        ]
        for check in checks:
            lines.append(f'    # [ ] {check}')
        lines.append(f'')

    # Export stub
    lines += [
        f'    # --- Export: {deliverable} ---',
        f'    # TODO: configure a QGIS print layout named "{map_id}_layout" then:',
        f'    # from qgis.core import QgsLayoutExporter',
        f'    # layout   = QgsProject.instance().layoutManager()',
        f'    #            .layoutByName("{map_id}_layout")',
        f'    # exporter = QgsLayoutExporter(layout)',
        f'    # exporter.exportToPdf("{map_id}_{short_name}.pdf",',
        f'    #     QgsLayoutExporter.PdfExportSettings())',
        f'',
        f'qgs.exitQgis()',
    ]

    return "\n".join(lines)


# ---------------------------------------------------------------------------
# Per-map ArcPy script generator
# ---------------------------------------------------------------------------

def generate_map_arcpy(m: dict, db_config: dict,
                       ops: list | None = None,
                       layer_info: dict | None = None) -> str:
    map_id      = m.get("map_id", "M??")
    title       = m.get("title") or m.get("short_name", "")
    short_name  = m.get("short_name", "layer")
    theme       = m.get("theme", "")
    subtheme    = m.get("subtheme", "")
    objective   = m.get("objective", "")
    questions   = m.get("key_questions", "")
    indicators  = m.get("key_indicators", "")
    scale       = m.get("study_scale", "")
    unit        = m.get("unit_of_analysis", "")
    sources     = m.get("data_sources", "")
    vintage     = m.get("data_vintage", "")
    processing  = m.get("processing_steps", "")
    symbology   = m.get("symbology_type", "")
    deliverable = m.get("deliverable_format", "")
    validation  = m.get("validation_checks", "")
    risks       = m.get("risks_limitations", "")
    status      = m.get("status", "")
    owner       = m.get("owner", "")
    layer_type  = m.get("spatial_layer_type", "Vector")

    var  = safe_var(short_name)
    host = db_config["host"]
    port = db_config["port"]
    db   = db_config["dbname"]
    user = db_config["user"]

    has_raster = "Raster" in layer_type

    lines = [
        f'"""',
        f'Map ID    : {map_id}',
        f'Title     : {title}',
        f'Theme     : {theme} > {subtheme}',
        f'Objective : {objective}',
        f'Questions : {questions}',
        f'Indicators: {indicators}',
        f'Scale     : {scale}  |  Unit: {unit}',
        f'Symbology : {symbology}',
        f'Sources   : {sources}  [{vintage}]',
        f'Processing: {processing}',
        f'Deliverable: {deliverable}',
        f'Validation: {validation}',
        f'Risks     : {risks}',
        f'Status    : {status}  |  Owner: {owner}',
        f'Generated : {datetime.now().strftime("%Y-%m-%d %H:%M")}',
        f'"""',
        f'',
        f'import arcpy',
        f'import os',
        f'import tempfile',
        f'',
        f'DB_HOST     = "{host}"',
        f'DB_PORT     = "{port}"',
        f'DB_NAME     = "{db}"',
        f'DB_USER     = "{user}"',
        f'DB_PASSWORD = os.environ["PGPASSWORD"]',
        f'',
        f'# --- SDE connection file ---',
        f'SDE_FOLDER = tempfile.gettempdir()',
        f'SDE_FILE   = os.path.join(SDE_FOLDER, f"{{DB_NAME}}.sde")',
        f'',
        f'if not os.path.exists(SDE_FILE):',
        f'    arcpy.management.CreateDatabaseConnection(',
        f'        out_folder_path=SDE_FOLDER,',
        f'        out_name=os.path.basename(SDE_FILE),',
        f'        database_platform="POSTGRESQL",',
        f'        instance=f"{{DB_HOST}},{{DB_PORT}}",',
        f'        account_authentication="DATABASE_AUTH",',
        f'        username=DB_USER,',
        f'        password=DB_PASSWORD,',
        f'        save_user_pass="SAVE_USERNAME",',
        f'        database=DB_NAME,',
        f'    )',
        f'    print(f"[OK] SDE connection created: {{SDE_FILE}}")',
        f'else:',
        f'    print(f"[OK] Reusing SDE connection: {{SDE_FILE}}")',
        f'',
        f'# {"=" * 66}',
        f'# {map_id} — {short_name}',
        f'# {title}',
        f'# {"=" * 66}',
        f'',
        f'fc_{var} = os.path.join(SDE_FILE, "public.{short_name}")',
        f'',
        f'if not arcpy.Exists(fc_{var}):',
        f'    print(f"[ERROR] \'{short_name}\' not found in SDE connection.")',
        f'else:',
        f'    desc_{var} = arcpy.Describe(fc_{var})',
        f'    print(f"[OK] {short_name}")',
        f'    print(f"  Geometry : {{desc_{var}.shapeType}}")',
        f'    print(f"  CRS      : {{desc_{var}.spatialReference.name}}")',
        f'    count_{var} = int(arcpy.management.GetCount(fc_{var})[0])',
        f'    print(f"  Rows: {{count_{var}}}")',
        f'',
        f'    # --- Add to ArcGIS Pro project ---',
        f'    # Use "CURRENT" when running inside the ArcGIS Pro Python console.',
        f'    # For standalone use, point APRX_PATH to your .aprx file.',
        f'    APRX_PATH = "CURRENT"  # TODO: or r"C:\\path\\to\\project.aprx"',
        f'    aprx   = arcpy.mp.ArcGISProject(APRX_PATH)',
        f'    mp_map = aprx.listMaps()[0]',
        f'    lyr_{var} = mp_map.addDataFromPath(fc_{var})',
        f'    print(f"  Added to map: {{mp_map.name}}")',
        f'',
    ]

    # Symbology — enrich field hint from schema if catalogue classification is blank
    m_sym = m
    if layer_info and not m.get("classification"):
        stype = (m.get("symbology_type") or "").lower()
        want_numeric = any(k in stype for k in ("choroplèthe", "dégradé", "gradué", "heatmap", "densité"))
        best = _best_field(layer_info, numeric=want_numeric)
        if best != "value":
            m_sym = {**m, "classification": best}
    lines.extend(_arcpy_symbology_block(var, m_sym))
    lines.append(f'')

    # Operation blocks (--op)
    if ops:
        columns = layer_info.get("columns", []) if layer_info else []
        lines.extend(_arcpy_op_blocks(var, short_name, columns, set(ops)))

    lines += [
        f'    aprx.save()',
        f'',
    ]

    # Raster note
    if has_raster:
        lines += [
            f'    # --- Raster component ---',
            f'    # This layer also has a raster component ({layer_type}).',
            f'    # TODO: load the raster via mp_map.addDataFromPath(raster_path).',
            f'    # raster_path = r"TODO: path to raster file"',
            f'    # lyr_{var}_raster = mp_map.addDataFromPath(raster_path)',
            f'',
        ]

    # Validation checks
    if validation:
        checks = [c.strip() for c in str(validation).split(",")]
        lines += [f'    # --- Validation checks ---']
        for check in checks:
            lines.append(f'    # [ ] {check}')
        lines.append(f'')

    # Export stub
    lines += [
        f'    # --- Export: {deliverable} ---',
        f'    # TODO: configure a layout named "{map_id}_layout" in your .aprx, then:',
        f'    # layout = aprx.listLayouts("{map_id}_layout")[0]',
        f'    # layout.exportToPDF("{map_id}_{short_name}.pdf")',
    ]

    return "\n".join(lines)


# ---------------------------------------------------------------------------
# Per-map QGIS project file generator
# ---------------------------------------------------------------------------

def generate_map_qgs(m: dict, db_config: dict,
                     ops: list | None = None,
                     layer_info: dict | None = None) -> str:
    """
    Generate a QGIS project file (.qgs) for a single catalogue map entry.

    ops is accepted for API consistency but silently ignored (.qgs has no
    operation blocks).  Returns XML — write to a .qgs file and open in QGIS.
    """
    short_name = m.get("short_name", "layer")

    if layer_info:
        geom         = layer_info.get("geometry",
                                      {"type": "GEOMETRY", "srid": 4326, "column": "geom"})
        primary_keys = layer_info.get("primary_keys", [])
        columns      = layer_info.get("columns", [])
    else:
        geom         = {"type": "GEOMETRY", "srid": 4326, "column": "geom"}
        primary_keys = []
        columns      = []

    single_layer_schema = {
        "database":    db_config.get("dbname", ""),
        "host":        db_config.get("host", ""),
        "layer_count": 1,
        "layers": [{
            "schema":         "public",
            "table":          short_name,
            "qualified_name": f"public.{short_name}",
            "geometry":       geom,
            "columns":        columns,
            "primary_keys":   primary_keys,
        }],
    }
    return generate_qgs(single_layer_schema, db_config)


# ---------------------------------------------------------------------------
# Per-map ArcGIS Python Toolbox generator
# ---------------------------------------------------------------------------

def generate_map_pyt(m: dict, db_config: dict,
                     ops: list | None = None,
                     layer_info: dict | None = None) -> str:
    """
    Generate an ArcGIS Python Toolbox (.pyt) for a single catalogue map entry.

    ops is accepted for API consistency but silently ignored (.pyt has no
    operation blocks).  Returns Python source — save as .pyt and open in
    ArcGIS Pro via Insert > Toolbox > Add Python Toolbox.
    """
    short_name = m.get("short_name", "layer")

    if layer_info:
        geom         = layer_info.get("geometry",
                                      {"type": "GEOMETRY", "srid": 4326, "column": "geom"})
        primary_keys = layer_info.get("primary_keys", [])
        columns      = layer_info.get("columns", [])
    else:
        geom         = {"type": "GEOMETRY", "srid": 4326, "column": "geom"}
        primary_keys = []
        columns      = []

    single_layer_schema = {
        "database":    db_config.get("dbname", ""),
        "host":        db_config.get("host", ""),
        "layer_count": 1,
        "layers": [{
            "schema":         "public",
            "table":          short_name,
            "qualified_name": f"public.{short_name}",
            "geometry":       geom,
            "columns":        columns,
            "primary_keys":   primary_keys,
        }],
    }
    return generate_pyt(single_layer_schema, db_config)


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

_NO_OP_PLATFORMS = {"qgs", "pyt"}   # ops silently ignored for these platforms
_EXT             = {"qgs": ".qgs", "pyt": ".pyt"}   # file extensions


def _build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(
        prog="gis-catalogue",
        description=(
            "Generate one script/project file per map from the Kensington catalogue."
        ),
    )
    p.add_argument("--input",  "-i", required=True, metavar="FILE",
                   help="Path to the catalogue .xlsx file.")
    p.add_argument("--output-dir", "-o", default="maps",
                   help="Directory to write generated files (default: ./maps/).")
    p.add_argument("--platform", "-p",
                   choices=["pyqgis", "arcpy", "qgs", "pyt"], default="pyqgis",
                   help="Target platform (default: pyqgis).")
    p.add_argument("--host",     default=os.environ.get("PGHOST",     "localhost"))
    p.add_argument("--port",     default=int(os.environ.get("PGPORT", 5432)), type=int)
    p.add_argument("--dbname",   default=os.environ.get("PGDATABASE", "my_gis_db"))
    p.add_argument("--user",     default=os.environ.get("PGUSER",     "postgres"))
    p.add_argument("--schema", "-s", metavar="FILE", default=None,
                   help="Schema JSON from 'gis-codegen --save-schema'. "
                        "Resolves field names and makes PGPASSWORD optional.")
    p.add_argument("--op", metavar="OPERATION", action="append", dest="operations",
                   choices=VALID_OPERATIONS,
                   help=f"Inject an operation block into every script "
                        f"({', '.join(VALID_OPERATIONS)}). Repeatable.")
    p.add_argument("--list",     action="store_true",
                   help="Print filtered maps and exit without writing files.")
    return p


def main() -> None:
    args = _build_parser().parse_args()

    maps = load_catalogue(args.input)
    print(f"[OK] Catalogue loaded: {len(maps)} Vector have/partial maps", file=sys.stderr)

    if args.list:
        print(f"\n  {'ID':<5}  {'STATUS':<8}  {'SHORT_NAME':<40}  SYMBOLOGY")
        print(f"  {'-'*5}  {'-'*8}  {'-'*40}  ---------")
        for m in maps:
            print(f"  {m['map_id']:<5}  {m['status']:<8}  {m['short_name']:<40}  {m['symbology_type']}")
        print()
        return

    # Load optional schema JSON (makes PGPASSWORD optional)
    schema_lookup: dict = {}
    if args.schema:
        schema_lookup = load_schema(args.schema)
        print(f"[OK] Schema loaded: {len(schema_lookup)} layers from {args.schema}",
              file=sys.stderr)
        # Backfill db_config from schema metadata when CLI args are at their defaults
        schema_raw = json.loads(Path(args.schema).read_text(encoding="utf-8"))
        if args.host == os.environ.get("PGHOST", "localhost"):
            args.host   = schema_raw.get("host",     args.host)
        if args.dbname == os.environ.get("PGDATABASE", "my_gis_db"):
            args.dbname = schema_raw.get("database", args.dbname)

    # qgs/pyt don't embed passwords so PGPASSWORD is not required for them
    if not os.environ.get("PGPASSWORD") and not args.schema \
            and args.platform not in _NO_OP_PLATFORMS:
        print("[ERROR] PGPASSWORD is not set. "
              "Either set it or supply --schema to skip the live DB requirement.",
              file=sys.stderr)
        sys.exit(1)

    db_config = {
        "host":   args.host,
        "port":   args.port,
        "dbname": args.dbname,
        "user":   args.user,
    }

    ops = args.operations or []
    if ops:
        if args.platform in _NO_OP_PLATFORMS:
            print(f"[WARN] --op is ignored for --platform {args.platform}",
                  file=sys.stderr)
        else:
            print(f"[OK] Operations: {', '.join(ops)}", file=sys.stderr)

    out_dir = Path(args.output_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    _generators = {
        "pyqgis": generate_map_pyqgis,
        "arcpy":  generate_map_arcpy,
        "qgs":    generate_map_qgs,
        "pyt":    generate_map_pyt,
    }
    generator = _generators[args.platform]
    ext       = _EXT.get(args.platform, ".py")

    for m in maps:
        fname      = f"{m['map_id']}_{m['short_name']}{ext}"
        fpath      = out_dir / fname
        layer_info = schema_lookup.get(m.get("short_name", ""))
        code       = generator(m, db_config, ops=ops, layer_info=layer_info)
        fpath.write_text(code, encoding="utf-8")
        resolved = layer_info is not None
        print(f"[OK] {fname}{' (schema enriched)' if resolved else ''}",
              file=sys.stderr)

    label = "files" if args.platform in _NO_OP_PLATFORMS else "scripts"
    print(f"\n[DONE] {len(maps)} {args.platform} {label} written to '{out_dir}/'",
          file=sys.stderr)


if __name__ == "__main__":
    main()
